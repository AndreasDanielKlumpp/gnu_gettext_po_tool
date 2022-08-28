#!/usr/bin/env python3

# -*- coding: utf-8 -*-

'''
gnu_gettext_po_tool.py

Table based GNU gettext po-file parser/writer

Copyright (c) 2015-2022 A.D.Klumpp

License: MIT

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

'''


_version = "5.03"

#gnu_gettext_po_tool v5.x
#aklumpp 2022

#Requirements:

#(Python 3.5.2)
#(Python 3.7.5)
#Python 3.9.2

#openpyxl 2.3.3
#jdcal 1.2
#et_xmlfile-1.0.1
#polib_a.py

#Licenses (Menu -> Write -> Licenses)
OPLH = "OPENPYXL LICENSE"
OPLT = "(http://openpyxl.readthedocs.org/en/latest/_modules/openpyxl/worksheet/header_footer.html) Copyright (c) 2010-2015 openpyxl. Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the 'Software'), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions: The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software. THE SOFTWARE IS PROVIDED 'AS IS', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE."

JDCALH = "jdcal 1.2 LICENSE"
JDCALT = '(https://pypi.python.org/pypi/jdcal) Copyright (c) 2011, Prasanth Nair. All rights reserved.\nRedistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:\n1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer.\n2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the documentation and/or other materials provided with the distribution.\nTHIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.'

ETXMLLH = "et_xmlfile 1.0.1 LICENSE"
ETXMLLT = "et_xmlfile is a low memory library for creating large XML files. It is based upon the xmlfile module from lxml <http://lxml.de/api.html#incremental-xml-generation>_ with the aim of allowing code to be developed that will work with both libraries. It was developed initially for the openpyxl project but is now a standalone module. The code was written by Elias Rabel as part of the Python Duesseldorf <http://pyddf.de>_ openpyxl sprint in September 2014. Version: 1.0.1. License: MIT. Home-page: https://bitbucket.org/openpyxl/et_xmlfile"

POLIBLH = "polib_a LICENSE"
POLIBLT = 'polib_a by A.Klumpp is a fork of polib 1.1.1 and has the same license.\n polib 1.1.1 License: Copyright (c) 2006-2015 David Jean Louis. \nPermission is hereby granted, free of charge, to any person obtaining a copyof this software and associated documentation files (the "Software"), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:\n The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software. \nTHE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.'

import re
import tkinter
from tkinter import *
from tkinter import filedialog
import tkinter.simpledialog
import tkinter.messagebox
from tkinter.filedialog import askopenfilename
from tkinter import Frame, Tk, BOTH, Text, Menu, END
import subprocess as sub
import pickle
import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from collections import Counter
import itertools
import sys
#po:
import polib_a
import os
import io
import ast
import time

#import xml.etree.ElementTree as ET
#sys.setdefaultencoding('utf-8')



class xlMain(Frame):



    def __init__(self, parent):
        Frame.__init__(self, parent)   

       
        self.parent = parent        
        self.initUI()

        frame = Frame(parent)        
        frame.pack()

  
   
        Label(parent, text="gnu_gettext_po_tool (ak) " + _version).pack()
        Label(parent, text="Table based GNU gettext po-file parser/writer").pack()
        Label(parent, text=" ").pack()
       

            


    def initUI(self):

        self.parent.title('gnu_gettext_po_tool')
        self.pack(fill=BOTH, expand=1)

        menubar = Menu(self.parent)
        self.parent.config(menu=menubar)

       
        fileMenu = Menu(menubar)
        fileMenu2 = Menu(menubar)
      
       
        menubar.add_cascade(label="Convert", menu=fileMenu)   

        fileMenu.add_command(label="po file -> table", command=self.ReadpoFileTool)
        fileMenu.add_command(label="Table -> po file", command=self.WritePoFile)
      
        fileMenu.add_command(label="LF: MS -> Unix", command=self.ConvertMSToUnix)
        fileMenu.add_command(label="LF: Unix -> MS", command=self.ConvertUnixToMS)

        menubar.add_cascade(label="Tools", menu=fileMenu2)       


        fileMenu2.add_command(label="LF of file (MS/Unix)", command=self.scanLF)
        fileMenu2.add_command(label="Table template", command=self.gettextTemplate)  
        fileMenu2.add_command(label="Licenses", command=self.Manual)
        


    def scanLF(self):


        ftypes = [('po files', '.po .pot')]
        dlg = filedialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

        if fl != '':
            filepath = fl
            filename = os.path.basename(filepath)
            print("Filename: " + str(filename))


        encoding_="ANSI"
        #with open(filepath, encoding=encoding_) as file:
        with open(filepath, encoding='utf-8') as file:
            #filedata = file.read()
            firstline = file.readline()
            print(firstline)
            firstnewline = file.newlines
            if firstnewline=="\r\n":
                print("MS")
            if firstnewline=="\n":
                print("UNIX")
                
 
        # "\r\n" (Windows), "\n" (Unix), "\r", None.
       
     



    def ConvertMSToUnix(self):
    

        WINDOWS_ = b'\r\n'
        UNIX_ = b'\n'


        year = datetime.datetime.now().year
        month = datetime.datetime.now().month
        hour = datetime.datetime.now().hour
        minute = datetime.datetime.now().minute
        day = datetime.datetime.now().day
        second = datetime.datetime.now().second


        filenameGL = str(day) + "_" + str(hour) + "_" + str(minute) + "_" + str(second) + '_unix.po'


        ftypes = [('po files', '.po .pot')]
        dlg = filedialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

        if fl != '':
            filepath = fl
            filename = os.path.basename(filepath)
            print("Filename: " + str(filename))


        with open(filepath, 'rb') as open_file:
            content = open_file.read()
            

        content = content.replace(WINDOWS_, UNIX_)


        with open(filenameGL, 'wb') as open_file:
            open_file.write(content)


        sys.stdout.write(filenameGL)
        
        sys.stdout.write(" ... Done.")
        sys.stdout.flush()        

        mainloop()   



    def ConvertUnixToMS(self):

        WINDOWS_ = b'\r\n'
        UNIX_ = b'\n'

        year = datetime.datetime.now().year
        month = datetime.datetime.now().month
        hour = datetime.datetime.now().hour
        minute = datetime.datetime.now().minute
        day = datetime.datetime.now().day
        second = datetime.datetime.now().second


        filenameGL = str(day) + "_" + str(hour) + "_" + str(minute) + "_" + str(second) + '_MS.po'


        ftypes = [('po files', '.po .pot')]
        dlg = filedialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

        if fl != '':
            filepath = fl
            filename = os.path.basename(filepath)
            print("Filename: " + str(filename))


        with open(filepath, 'rb') as open_file:
            content = open_file.read()
            

        content = content.replace(UNIX_, WINDOWS_)

        with open(filenameGL, 'wb') as open_file:
            open_file.write(content)


        sys.stdout.write(filenameGL)
        
        sys.stdout.write(" ... Done.")
        sys.stdout.flush()        

        mainloop()   

    def gettextTemplate(self):

        print("Saving gnu gettext DB table...")

        year = datetime.datetime.now().year
        month = datetime.datetime.now().month
        hour = datetime.datetime.now().hour
        minute = datetime.datetime.now().minute
        day = datetime.datetime.now().day
        second = datetime.datetime.now().second

        filenameGL = str(day) + "_" + str(hour) + "_" + str(minute) + "_" + str(second) + '_po_template.xlsx'

        wbgl = Workbook()
        
        wsgl1 = wbgl.active
        wsgl1.title = "msg"
        wsgl2 = wbgl.create_sheet('comment')
        wsgl3 = wbgl.create_sheet('metadata')
        #wsgl4 = wbgl.create_sheet('occurrences')
        #wsgl5 = wbgl.create_sheet('flags')


        wsgl3.cell(row= 1, column=1).value = "File"
        #wsgl3.cell(row= 2, column=1).value = str(filename)        
      
        #mdatah = pofile.header
        #print("mdatah: " + str(mdatah))
        wsgl3.cell(row= 1, column=2).value = "Header"
        #wsgl3.cell(row= 2, column=2).value = str(mdatah)

       
        #print("mdatab: " + str(mdatab))
        wsgl3.cell(row= 1, column=3).value = "Metadata"
        #wsgl3.cell(row= 1, column=3).value = "Metadata key"
        #wsgl3.cell(row= 1, column=4).value = "Metadata value"
        #wsgl3.cell(row= 2, column=3).value = str(mdatab)

        wsgl3.cell(row= 1, column=4).value = "Settings"
        wsgl3.cell(row= 1, column=5).value = "Value"
        wsgl3.cell(row= 2, column=4).value = "Encoding:"
        wsgl3.cell(row= 2, column=5).value = "utf-8"
        wsgl3.cell(row= 1, column=6).value = "Example"
        wsgl3.cell(row= 2, column=6).value = "utf-8, utf-8-sig, utf16, ANSI (auto selected)"
        wsgl3.cell(row= 3, column=4).value = "Multiline Strings:"
        wsgl3.cell(row= 3, column=5).value = "No"
        wsgl3.cell(row= 3, column=6).value = "No/Yes"
        wsgl3.cell(row= 4, column=4).value = "Line breaks (OS):"
        wsgl3.cell(row= 4, column=5).value = "MS"
        wsgl3.cell(row= 4, column=6).value = "MS/UNIX (auto selected)"
        wsgl3.cell(row= 5, column=4).value = "Line breaks (String):"
        wsgl3.cell(row= 5, column=5).value = "MS"
        wsgl3.cell(row= 5, column=6).value = "MS/UNIX"
        wsgl3.cell(row= 6, column=4).value = "All Selectors:"
        wsgl3.cell(row= 6, column=5).value = "Yes"
        wsgl3.cell(row= 6, column=6).value = "No/Yes"
                
        wsgl1.cell(row= 1, column=1).value = "String_ID"
        wsgl2.cell(row= 1, column=1).value = "String_ID"
        wsgl2.cell(row= 1, column=2).value = "comment 1"
        wsgl2.cell(row= 1, column=3).value = "comment 2..."
        
        wsgl1.cell(row= 1, column=2).value = "filename"
        #wsgl1.cell(row= 1, column=3).value = "comment (#.)"
        wsgl1.cell(row= 1, column=3).value = "tcomment (#)"
        wsgl1.cell(row= 1, column=4).value = "occurrences (#:)"
        wsgl1.cell(row= 1, column=5).value = "flags (#,)"
        wsgl1.cell(row= 1, column=6).value = "msgctxt"
        wsgl1.cell(row= 1, column=7).value = "previous_msgid (#|)"
        wsgl1.cell(row= 1, column=8).value = "previous_msgctxt"
        wsgl1.cell(row= 1, column=9).value = "previous_msgid_plural"
        wsgl1.cell(row= 1, column=10).value = "linenum" 
        wsgl1.cell(row= 1, column=11).value = "msgid"
        wsgl1.cell(row= 1, column=12).value = "msgid_plural"        
        wsgl1.cell(row= 1, column=13).value = "msgstr"
       
      

        
        wbgl.save(filename = filenameGL)

        sys.stdout.write(filenameGL)
        sys.stdout.write(" ... Done.")
        sys.stdout.flush()

        mainloop()        

    def ReadpoFileTool(self):

        '''
        Examples of gettext 'Selectors':

        #  gettext translator-comment-1
        #  gettext translator-comment-2
        #  gettext translator-comment-3...
        #. gettext comment-line-1
        #. gettext comment-line-2
        #. gettext comment-line-3
        #. gettext comment-line-4
        #. gettext comment-line-5
        #. gettext comment-line-6...
        #: gettext occurrences
        #, gettext flags
        #| msgid previous-untranslated-string
        msgctxt c(on)text...
        msgid untranslated-string
        msgid_plural untranslated-string
        msgstr translated-string
        msgstr[0] translated-string_plural
        msgstr[1] translated-string_plural
        msgstr[2] translated-string_plural
        msgstr[3] translated-string_plural
        msgstr[4] translated-string_plural
        msgstr[5] translated-string_plural
        ...


        '''

        #disable multiline msgstr/id
        #if empty it should not be none
        #check if selecor is/which selecrors are present

        #if unescaped -> escape instead of error
        
        #not used:
        #translated_entries()/untranslated_entries()
        #percent_translated()
        #obsolete_entries()
        #fuzzy_entries()
        #check_for_duplicates
        

        #todo
        #occurances dictionary 
        
        #read/write multiple po files

        #polib.unescape(string)
        

        print("ReadpoFileTool")

        year = datetime.datetime.now().year
        month = datetime.datetime.now().month
        hour = datetime.datetime.now().hour
        minute = datetime.datetime.now().minute
        day = datetime.datetime.now().day
        second = datetime.datetime.now().second



        filenameGL = str(day) + "_" + str(hour) + "_" + str(minute) + "_" + str(second) + '_po_tool.xlsx'

        wbgl = Workbook()
        
        wsgl1 = wbgl.active
        wsgl1.title = "msg"
        wsgl2 = wbgl.create_sheet('comment')
        wsgl3 = wbgl.create_sheet('metadata')
        #wsgl4 = wbgl.create_sheet('occurrences')
        #wsgl5 = wbgl.create_sheet('flags')




        ftypes = [('po files', '.po .pot')]
        dlg = filedialog.Open(self, filetypes = ftypes)
        fl = dlg.show()

        if fl != '':
            filepath = fl
            filename = os.path.basename(filepath)
            print("Filename: " + str(filename))

        #scan line feed
        with open(filepath, encoding='utf-8') as file:
            #filedata = file.read()
            firstline = file.readline()
            #print(firstline)
            firstnewline = file.newlines
            if firstnewline=="\r\n":
                Linef="MS"
                print("OS: MS")
            if firstnewline=="\n":
                Linef="UNIX"
                print("OS: UNIX")
        
            

        #pofile = polib.pofile('/.../py_tool/example_02.po')
        pofile = polib_a.pofile(filepath)
        
        pencodeing = polib_a.detect_encoding(filepath,FALSE)
        print("Encoding: " + str(pencodeing))
        if pencodeing=="UTF-8":
            pencodeing="utf-8"

        #mdata = pofile.metadata.get('Language-Team')
        #print("mdata: " + str(mdata))


        wsgl3.cell(row= 1, column=1).value = "File"
        #wsgl3.cell(row= 2, column=1).value = str(filename)        
      
        mdatah = pofile.header
        #print("mdatah: " + str(mdatah))
        wsgl3.cell(row= 1, column=2).value = "Header"
        #wsgl3.cell(row= 2, column=2).value = str(mdatah)

        full_header = str(mdatah)
            
        sheader = io.StringIO(full_header)

        headerline = 0
        for hline in sheader:            
            headerline = headerline +1
            print("hline: " + str(headerline) + " : " + str(hline))
            wsgl3.cell(row= headerline+1, column=2).value = str(hline)
            wsgl3.cell(row= headerline+1, column=1).value = str(filename) 

        mdatab = pofile.metadata
        #print("mdatab: " + str(mdatab))
        wsgl3.cell(row= 1, column=3).value = "Metadata"
        #wsgl3.cell(row= 1, column=3).value = "Metadata key"
        #wsgl3.cell(row= 1, column=4).value = "Metadata value"
        #wsgl3.cell(row= 2, column=3).value = str(mdatab)

        wsgl3.cell(row= 1, column=4).value = "Settings"
        wsgl3.cell(row= 1, column=5).value = "Value"
        wsgl3.cell(row= 2, column=4).value = "Encoding:"
        wsgl3.cell(row= 2, column=5).value = pencodeing
        wsgl3.cell(row= 1, column=6).value = "Example"
        wsgl3.cell(row= 2, column=6).value = "utf-8, utf-8-sig, utf16, ANSI (auto selected)"
        wsgl3.cell(row= 3, column=4).value = "Multiline Strings:"
        wsgl3.cell(row= 3, column=5).value = "No"
        wsgl3.cell(row= 3, column=6).value = "No/Yes"
        wsgl3.cell(row= 4, column=4).value = "Line breaks (OS):"
        wsgl3.cell(row= 4, column=5).value = Linef
        wsgl3.cell(row= 4, column=6).value = "MS/UNIX (auto selected)"
        wsgl3.cell(row= 5, column=4).value = "Line breaks (String):"
        wsgl3.cell(row= 5, column=5).value = "MS"
        wsgl3.cell(row= 5, column=6).value = "MS/UNIX"
        wsgl3.cell(row= 6, column=4).value = "All Selectors:"
        wsgl3.cell(row= 6, column=5).value = "Yes"
        wsgl3.cell(row= 6, column=6).value = "No/Yes"

        
        #mdatab_adk = pofile.metadata_adk
        #print("mdatab_adk: " + str(mdatab_adk))

        mdatab_adk_arr = pofile.metadata_adk_arr
        #print("mdatab_adk_arr: " + str(mdatab_adk_arr))

        mdatab_adk_arrlength = len(mdatab_adk_arr)

        metaline = 0
        for mvalue in range(len(mdatab_adk_arr)):
            metaline = metaline +1
            print("mvalue: " + str(mdatab_adk_arr[mvalue]))         
            wsgl3.cell(row= metaline+1, column=3).value = str(mdatab_adk_arr[mvalue])
            wsgl3.cell(row= metaline+1, column=1).value = str(filename)
            
        '''    
        metaline = 0
        for key, value in mdatab.items():
            metaline = metaline +1
            print("key: " + str(metaline) + " : " + str(key))
            print("value: " + str(metaline) + " : " + str(value))         
            wsgl3.cell(row= metaline+1, column=3).value = str(key)
            wsgl3.cell(row= metaline+1, column=4).value = str(value)
            wsgl3.cell(row= metaline+1, column=1).value = str(filename)
        '''
        
        entr = int(0)
        for entry in pofile:
            entr = entr + int(1)
            #print("Progress (py_tool_ID): " + str(entr))
            
            #print("msgid: " + str(entry.msgid))
            #print("msgstr: " + str(entry.msgstr))
            #print("msgctxt: " + str(entry.msgctxt))
            
            #print("occurrences: " + str(entry.occurrences))
            #reference
            #po:
            #: abc
            #output:
            #-> occurrences: [('abc', '')]
            
            #print("comment: " + str(entry.comment))
            #po:
            #. Line 1
            #. Line 2
            #. Line 3
            #output:
            #-> comment: Line 1
            #-> Line 2
            #-> Line 3
            
            #print("tcomment: " + str(entry.tcomment))
            #po:
            # yyyWikipedia says that
            #output:
            #-> tcomment: xxxWikipedia says that...
            
            #print("flags: " + str(entry.flags))
            #po:
            #, xxxc-format
            #output:
            #-> flags: ['xxxc-format']




            
            #merge entries of flags and occurrences array (flags: 1 dim., occurrences: 2 dim.)
            
       
            flen = len(entry.flags)
            #print(str(flen))
            merge_fl = []
            for fpart in range(int(0),int(flen)):
                fl_pt = entry.flags[fpart]
                #print(str(entry.flags[opart][0]))
                merge_fl.append(str(fl_pt))
            #print(str(merge_oc))
            conc_fl = ''.join(merge_fl)
            #print(str(conc_fl))            

            '''
            olen = len(entry.occurrences)
            #print(str(olen))
            merge_oc = []
            for opart in range(int(0),int(olen)):
                oc_pt = entry.occurrences[opart][0]
                #print(str(entry.occurrences[opart][0]))
                merge_oc.append(str(oc_pt))
            #print(str(merge_oc))
            conc_occur = ' '.join(merge_oc)
            #print(str(conc_occur))
            '''


            plen = len(entry.msgstr_plural)
            #print(str(olen))

            p_header = 0
            for ppart in range(int(0),int(plen)):
                p_header = p_header + 1
                p_pt = entry.msgstr_plural[ppart]
                #print(str(p_pt))
                wsgl1.cell(row= 1, column=13+p_header).value = str(p_header) + ". msgstr_plural"
                wsgl1.cell(row= entr+1, column=13+p_header).value = str(p_pt)

                

            #one cell per comment:
            
            full_comment = str(entry.comment)
            
            s = io.StringIO(full_comment)
            #print(s)                

            
            strline = 0
            for line in s:
                strline = strline +1
                wsgl2.cell(row= 1, column=1+strline).value = str(strline) + ". comment (#.)"
                wsgl2.cell(row= entr+1, column=1+strline).value = str(line)

               
            escapeswitch = 1
            unescapeswitch = 0

            if(escapeswitch==1):
                msgstr_ = polib_a.escape(entry.msgstr)
                #print("esc.: " + str(msgstr_))
            else:
                msgstr_ = entry.msgstr
                
            if(unescapeswitch==1):
                msgstr_ = polib_a.unescape(entry.msgstr)
                #print("unesc.: " + str(msgstr_))
            else:
                msgstr_ = entry.msgstr
            
            if(escapeswitch==1):
                msgid_ = polib_a.escape(entry.msgid)
            else:
                msgid_ = entry.msgid
                
            if(unescapeswitch==1):
                msgid_ = polib_a.unescape(entry.msgid)
            else:
                msgid_ = entry.msgid
            

                    
            py_tool_ID = entr

            wsgl1.cell(row= entr+1, column=1).value = str(py_tool_ID)
            wsgl2.cell(row= entr+1, column=1).value = str(py_tool_ID)
            
            wsgl1.cell(row= entr+1, column=2).value = str(filename)
            #wsgl1.cell(row= entr+1, column=3).value = str(entry.comment)
            wsgl1.cell(row= entr+1, column=3).value = str(entry.tcomment)
            
            #merge array of occurences:
            wsgl1.cell(row= entr+1, column=4).value = str(entry.occurrences)
            #wsgl1.cell(row= entr+1, column=4).value = str(conc_occur)

            #merge array of occurences:
            #wsgl1.cell(row= entr+1, column=5).value = str(entry.flags)
            wsgl1.cell(row= entr+1, column=5).value = str(conc_fl)
            
            wsgl1.cell(row= entr+1, column=6).value = str(entry.msgctxt)
            wsgl1.cell(row= entr+1, column=7).value = str(entry.previous_msgid)
            wsgl1.cell(row= entr+1, column=8).value = str(entry.previous_msgctxt)
            wsgl1.cell(row= entr+1, column=9).value = str(entry.previous_msgid_plural)
            wsgl1.cell(row= entr+1, column=10).value = str(entry.linenum)
            
            #wsgl1.cell(row= entr+1, column=11).value = str(entry.msgid)
            wsgl1.cell(row= entr+1, column=11).value = str(msgid_)            
            wsgl1.cell(row= entr+1, column=12).value = str(entry.msgid_plural)
            #wsgl1.cell(row= entr+1, column=13).value = str(entry.msgstr)
            wsgl1.cell(row= entr+1, column=13).value = str(msgstr_)
          

            
            
            #wsgl1.cell(row= entr+1, column=15).value = str(entry.msgstr_plural)



        wsgl1.cell(row= 1, column=1).value = "String_ID"
        wsgl2.cell(row= 1, column=1).value = "String_ID"
        
        wsgl1.cell(row= 1, column=2).value = "filename"
        #wsgl1.cell(row= 1, column=3).value = "comment (#.)"
        wsgl1.cell(row= 1, column=3).value = "tcomment (#)"
        wsgl1.cell(row= 1, column=4).value = "occurrences (#:)"
        wsgl1.cell(row= 1, column=5).value = "flags (#,)"
        wsgl1.cell(row= 1, column=6).value = "msgctxt"
        wsgl1.cell(row= 1, column=7).value = "previous_msgid (#|)"
        wsgl1.cell(row= 1, column=8).value = "previous_msgctxt"
        wsgl1.cell(row= 1, column=9).value = "previous_msgid_plural"
        wsgl1.cell(row= 1, column=10).value = "linenum" 
        wsgl1.cell(row= 1, column=11).value = "msgid"
        wsgl1.cell(row= 1, column=12).value = "msgid_plural"        
        wsgl1.cell(row= 1, column=13).value = "msgstr"
       
      

        
        wbgl.save(filename = filenameGL)

        sys.stdout.write(filenameGL)
        sys.stdout.write(" ... Done.")
        sys.stdout.flush()

        mainloop()



    def WritePoFile(self):



        AllSelectors=0

        year = datetime.datetime.now().year
        month = datetime.datetime.now().month
        hour = datetime.datetime.now().hour
        minute = datetime.datetime.now().minute
        day = datetime.datetime.now().day
        second = datetime.datetime.now().second



        filenameGL = str(day) + "_" + str(hour) + "_" + str(minute) + "_" + str(second) + '.po'
     
        filenameGL_lit = filenameGL + "_lit" + '.po'

        #escapeswitch = 1: everthing in one line with escape commands instead of real line breaks
        #escapeswitch = 0: multi-line
        #ms_switch: 1: \r\n - 0: \n
        #literalonly: don't save the linebreak with double escape 1: \n - 0: \\n
       
        #encoding_: read from import table (UTF-8, utf-8-sig, utf16, ANSI...)

        #in table
        escapeswitch = 1

        #OS LF: unix (0), ms (1)
        ms_switch = 1

        #string contains \n\r (1) or \n (0) - not depending on global (OS) LF
        str_contains_n_r = 1

        #not in table
        literalonly = 1
        unescapeswitch = 0        
       


        #arr_msgstrs = []
        #arr_msgstrs.append("pl_1")
        #arr_msgstrs.append("pl_2")
        #print("plural arr: " + str(arr_msgstrs))


        po = polib_a.POFile()

        #write header from header info tab

       
        
        ftypes = [('Excel files', '.xlsm .xlsx')]
        dlg = filedialog.Open(self, filetypes = ftypes)
        fl = dlg.show()



        if fl != '':
            filename = fl
            book = load_workbook(filename)
            sheetgl = book.worksheets[0]
            sheetg2 = book.worksheets[1]
            sheetg3 = book.worksheets[2]

        print("Source: " + str(filename))               

        print("Writing .po file...")


        encoding_ = sheetg3.cell(row = 2,column = 5).value

        if encoding_=="" or encoding_==None:
             encoding_ = "UTF-8"
             
        print("Encoding: " + str(encoding_))
        
        escapeswitch_ = sheetg3.cell(row = 3,column = 5).value

        if escapeswitch_=="" or escapeswitch_==None or escapeswitch_=="No":
             escapeswitch = 1
             print("Multi-line strings: Off")

        if escapeswitch_=="Yes":
             escapeswitch = 0
             print("Multi-line strings: On")
#

        
        ms_switch_ = sheetg3.cell(row = 4,column = 5).value

        if ms_switch_=="" or ms_switch_==None or ms_switch_=="MS":
             ms_switch = 1
             print("Line breaks (OS): MS")

        if ms_switch_=="UNIX":
             ms_switch = 0
             print("Line breaks (OS): UNIX")
        
#                   
        str_contains_n_r_ = sheetg3.cell(row = 5,column = 5).value

        if str_contains_n_r_=="" or str_contains_n_r_==None or str_contains_n_r_=="MS":
             str_contains_n_r = 1
             print("Line breaks (String): MS")

        if str_contains_n_r_=="UNIX":
             str_contains_n_r = 0
             print("Line breaks (String): UNIX")

        all_selectors_switch = sheetg3.cell(row = 6,column = 5).value

        if all_selectors_switch=="" or all_selectors_switch==None or all_selectors_switch=="No":
             AllSelectors = 0
             print("All Selectors: Off")

        if all_selectors_switch =="Yes":
             AllSelectors = 1
             print("All Selectors: On")


        #po.header = sheetg3.cell(row = 2,column = 2).value
        #string_metadata = sheetg3.cell(row = 2,column = 3).value
        #print("string_metadata : " + str(string_metadata))
        
        #po.metadata = ast.literal_eval(string_metadata)
        
        #po.header = "line1\nline2\nline3"
        #po.metadata_adk_arr = ['Project-Id-Version: xyz', 'POT-Creation-Date: 2021-05-11 10:14', 'PO-Revision-Date: 2021-06-17 10:52+0300', 'Language-Team: French', 'Language: fr']

        metadata_arr = []
        row_countgl = sheetgl.max_row+1
        col_countgl = sheetgl.max_column+1
        col_countg2 = sheetg2.max_column+1

        row_countg3 = sheetg3.max_row+1
        
        firstrowtoread=2


              

        for nheader in range(int(firstrowtoread), int(row_countg3)):                  

            read_header = sheetg3.cell(row = nheader,column = 2).value

            if read_header != None:              
                po.header += read_header
            else:
                print("Empty header cell: " + str(nheader))
                #po.header = ""
                


        notemptyswitch = 0
        for nmetadata in range(int(firstrowtoread), int(row_countg3)):                  

            read_metadata = sheetg3.cell(row = nmetadata,column = 3).value

            if read_metadata != None:
                notemptyswitch = 1
                metadata_arr.append(read_metadata)
            else:
                print("Empty metadata cell: " + str(nmetadata))

        if (notemptyswitch == 1):
            po.metadata_adk_arr = metadata_arr
            


        for n in range(int(firstrowtoread), int(row_countgl)):

            #print(n)
            pl_1_switch=0
            pl_2_switch=0
            pl_3_switch=0
            pl_4_switch=0
            pl_5_switch=0
            pl_6_switch=0
            pl_7_switch=0

            read_msgid = sheetgl.cell(row = n,column = 11).value
            read_msgstr = sheetgl.cell(row = n,column = 13).value
            read_msgctxt = sheetgl.cell(row = n,column = 6).value
            read_tcomment = sheetgl.cell(row = n,column = 3).value
            read_occurrences = sheetgl.cell(row = n,column = 4).value

            if AllSelectors==1:
                read_previous_msgid = sheetgl.cell(row = n,column = 7).value              
                read_flags = sheetgl.cell(row = n,column = 5).value
                read_previous_msgctxt = sheetgl.cell(row = n,column = 8).value
                read_previous_msgid_plural = sheetgl.cell(row = n,column = 9).value
                read_msgid_plural = sheetgl.cell(row = n,column = 12).value

                
                pl_1_ = sheetgl.cell(row = n,column = 14).value
                if pl_1_ is None or pl_1_=="" or pl_1_=="None":
                    #print("pl_1 is None")
                    pl_1_switch=0
                else:
             
                    pl_1_switch=1
                    pl_1 = sheetgl.cell(row = n,column = 14).value
                    print(pl_1)
                
                
                pl_2_ = sheetgl.cell(row = n,column = 15).value
                if pl_2_ is None  or pl_2_=="" or pl_2_=="None":
                    #print("pl_2 is None")
                    pl_2_switch=0

                else:

                    pl_2_switch=1
                    pl_2 = sheetgl.cell(row = n,column = 15).value
                    print(pl_2)
               

                pl_3_ = sheetgl.cell(row = n,column = 16).value
                if pl_3_ is None  or pl_3_=="" or pl_3_=="None":
                    #print("pl_3 is None")
                    vpl_3_switch=0

                else:

                    pl_3_switch=1
                    pl_3 = sheetgl.cell(row = n,column = 16).value
                    print(pl_3)

                pl_4_ = sheetgl.cell(row = n,column = 17).value
                if pl_4_ is None  or pl_4_=="" or pl_4_=="None":
                    #print("pl_4 is None")
                    pl_4_switch=0

                else:

                    pl_4_switch=1
                    pl_4 = sheetgl.cell(row = n,column = 17).value
                    print(pl_4)

                pl_5_ = sheetgl.cell(row = n,column = 18).value
                if pl_5_ is None  or pl_5_=="" or pl_5_=="None":
                    #print("pl_5 is None")
                    pl_5_switch=0

                else:

                    pl_5_switch=1
                    pl_5 = sheetgl.cell(row = n,column = 18).value
                    print(pl_5)

                    
                pl_6_ = sheetgl.cell(row = n,column = 19).value
                if pl_6_ is None  or pl_6_=="" or pl_6_=="None":
                    #print("pl_6 is None")
                    pl_6_switch=0

                else:

                    pl_6_switch=1
                    pl_6 = sheetgl.cell(row = n,column = 19).value
                    print(pl_6)


                pl_7_ = sheetgl.cell(row = n,column = 20).value
                if pl_7_ is None  or pl_7_=="" or pl_7_=="None":
                    #print("pl_7 is None")
                    pl_7_switch=0

                else:

                    pl_7_switch=1
                    pl_7 = sheetgl.cell(row = n,column = 20).value
                    print(pl_7)
                    
         
            #unlimited comment cells in tab2:

            read_comment_ = ""
            read_comment = ""

            for c in range(int(2), int(col_countg2)): 

            
                read_comment1 = sheetg2.cell(row = n,column = c).value            
           

                
                if (read_comment1 != "None" and read_comment1 != "" and read_comment1 != None):
                    read_comment1_ = read_comment1.replace('\n', '').replace('\r', '')
                    if(c==2):
                        read_comment_ =  read_comment1_
                    else:
                        read_comment_ +=  "\n" + read_comment1_
                              
                if read_comment_ != "":
                    read_comment = read_comment_
                    



            #test with empty cells
            #escape line break msgstr      
            
            if read_msgstr != None:
           
                #read_msgstr_ = read_msgstr
                #read_msgstr_ = read_msgstr.replace('\\xa0', r'\xa0')
                #excluding non breaking spaces for now
                read_msgstr_ = read_msgstr.replace('\xa0', ' ')
             

                
           
         
                if(escapeswitch==1):
                    #print(str(n) + " : " + str(read_msgstr_))
                    #

                    full_cell = str(read_msgstr_)
                
                    sc = io.StringIO(full_cell)
                    #print(s)                

                    lines_arr = []
                    strlinec = 0
                    for line in sc:
                        strlinec = strlinec +1
                        #print("line: " + str(line))
                        lines_arr.append(line)
                    #print(lines_arr)
                        
                  
                    #read_msgstr_esc = polib_a.escape(read_msgstr)
                    
                    #read_msgstr_esc = repr(read_msgstr)
                    #print(read_msgstr_esc)
                    #read_msgstr_ = read_msgstr_esc

                    jlines = ''.join(lines_arr)
                    
                    jlinesr = repr(jlines)[1:-1]
                   
                    #if "\\n" in jlinesr:
                        #print(jlinesr)
                        #print("--------------------------------------")
                    #arr_linesr.append(jlinesr)
                    #print(arr_linesr) 
                    #print(repr(jlinesr))

                    #read_msgstr_esc = jlinesr.replace('\\r\\n', r'\r\n')

                    #read_msgstr_ = jlinesr
                    '''
                    if not '\r\n' in jlinesr:
                        read_msgstr_esc = jlinesr.replace('\\n', r'\n')
                    #0803
                    else:
                        read_msgstr_esc = read_msgstr_esc.replace('\\n', r'\r\n')
                    '''
                        
                    #read_msgstr_ = read_msgstr_esc
                    
                                                   
                    #if "\\n" in read_msgstr_:
                        #print(read_msgstr_)
                        #print("--------------------------------------")
                    
                    if str_contains_n_r==1:
                        read_msgstr__ = jlinesr.replace('\\n', r'\r\n')
                        read_msgstr_ = read_msgstr__.replace("\\'", "'")

                        #read_msgstr_ = read_msgstr_ms.replace('\\r\\n', '\r\n')
                        #read_msgstr_ = read_msgstr_ms.encode("unicode_escape").decode("utf-8")
                        
                        #read_msgstr_ = read_msgstr_ms
                        
                    if str_contains_n_r==0:
                        read_msgstr__ = jlinesr.replace('\\n', r'\n')
                        read_msgstr_ = read_msgstr__.replace("\\'", "'")
                    
                    
                if(unescapeswitch==1):
                    read_msgstr_ = polib_a.unescape(read_msgstr)
                    
            else:
                #print("Empty msgstr cell: " + str(n))
                read_msgstr_ = ""


            #escape line break msgid
                
            if read_msgid != None:                
                read_msgid_ = read_msgid
                
                if(escapeswitch==1):
                    #read_msgid_esc = polib_a.escape(read_msgid)
                    #read_msgid_ = read_msgid_esc


                    full_cell = str(read_msgid_)
                
                    sc = io.StringIO(full_cell)
                    #print(s)                

                    lines_arr = []
                    strlinec = 0
                    for line in sc:
                        strlinec = strlinec +1
                      
                        lines_arr.append(line)
      

                    jlines = ''.join(lines_arr)
                    
                    jlinesr = repr(jlines)[1:-1]
           
                    #read_msgid_ = jlinesr
                    #read_msgid_esc = jlinesr.replace(r'\\n', r'\n')
                    #read_msgid_ = read_msgid_esc
                    
                    if str_contains_n_r==0:
                        read_msgid__ = jlinesr.replace('\\n', r'\n')
                        read_msgid_ = read_msgid__.replace("\\'", "'")
                        
                    

                    elif str_contains_n_r==1:
                        read_msgid__ = jlinesr.replace('\\n', r'\r\n')
                        read_msgid_ = read_msgid__.replace("\\'", "'")

                    

                        #read_msgid_ = read_msgid_ms.replace('\\r\\n', '\r\n')



                    
                if(unescapeswitch==1):
                    read_msgid_ = polib_a.unescape(read_msgid)
                    
            else:
                #print("Empty msgid cell: " + str(n))
                read_msgid_ = ""                

            if read_occurrences != None:                
                read_occurrences_ = read_occurrences
                #print("occ : " + str(read_occurrences_))
            else:
                #print("Empty occurrences cell: " + str(n))
                read_occurrences_ = ""
              
            if read_tcomment != None:                
                read_tcomment_ = read_tcomment
            else:
                #print("Empty tcomment cell: " + str(n))
                read_tcomment_ = ""

            #maybe add a switch to print none msgctxt cells either as none or empty
            if read_msgctxt == "None":
                read_msgctxt_ = ""
            elif read_msgctxt != None:                
                read_msgctxt_ = read_msgctxt
            else:
                #print("Empty msgctxt cell: " + str(n))
                read_msgctxt_ = ""





            '''
   
            wsgl1.cell(row= 1, column=4).value = "occurrences (#:)"

            wsgl1.cell(row= 1, column=5).value = "flags (#,)"
          
            wsgl1.cell(row= 1, column=7).value = "previous_msgid (#|)"
            wsgl1.cell(row= 1, column=8).value = "previous_msgctxt"
            wsgl1.cell(row= 1, column=9).value = "previous_msgid_plural"
            wsgl1.cell(row= 1, column=10).value = "linenum" 
           
            wsgl1.cell(row= 1, column=12).value = "msgid_plural"        
       

            '''

      
   

   
            c=","
            e1="comment=read_comment"
            e2="occurrences=read_occurrences_"
            #e2="occurrences=[(read_occurrences_, '')]"
            e3="tcomment=read_tcomment_"
            e4="msgctxt=read_msgctxt_"
            e5="msgid=read_msgid_"
            e6="msgstr=read_msgstr_"

            e_=e1+c+e2+c+e3+c+e4+c+e5+c+e6
            
           


            if AllSelectors==1:
                e7="previous_msgid=read_previous_msgid"               
                e9="flags=read_flags"
                e10="previous_msgctxt=read_previous_msgctxt"
                e11="previous_msgid_plural=read_previous_msgid_plural"
                e12="msgid_plural=read_msgid_plural"

                #e13="msgstr_plural=read_msgstr_plural"


                if pl_1_switch==1 and pl_2_switch==1 and pl_3_switch==0:
                    print("pl_1/2")
                    e13='msgstr_plural=[pl_1, pl_2]'
                    e_+=c+e7+c+e9+c+e10+c+e11+c+e12+c+e13
                    
                elif pl_1_switch==1 and pl_2_switch==1 and pl_3_switch==1 and pl_4_switch==0:
                    print("pl_1/2/3")
                    e13='msgstr_plural=[pl_1, pl_2, pl_3]'
                    e_+=c+e7+c+e9+c+e10+c+e11+c+e12+c+e13

                elif pl_1_switch==1 and pl_2_switch==1 and pl_3_switch==1 and pl_4_switch==1 and pl_5_switch==0:
                    print("pl_1/2/3/4")
                    e13='msgstr_plural=[pl_1, pl_2, pl_3, pl_4]'
                    e_+=c+e7+c+e9+c+e10+c+e11+c+e12+c+e13

                elif pl_1_switch==1 and pl_2_switch==1 and pl_3_switch==1 and pl_4_switch==1 and pl_5_switch==1 and pl_6_switch==0:
                    print("pl_1/2/3/4/5")
                    e13='msgstr_plural=[pl_1, pl_2, pl_3, pl_4, pl_5]'
                    e_+=c+e7+c+e9+c+e10+c+e11+c+e12+c+e13

                elif pl_1_switch==1 and pl_2_switch==1 and pl_3_switch==1 and pl_4_switch==1 and pl_5_switch==1 and pl_6_switch==1 and pl_7_switch==0:
                    print("pl_1/2/3/4/5/6")
                    e13='msgstr_plural=[pl_1, pl_2, pl_3, pl_4, pl_5, pl_6]'
                    e_+=c+e7+c+e9+c+e10+c+e11+c+e12+c+e13
              
                elif pl_1_switch==1 and pl_2_switch==1 and pl_3_switch==1 and pl_4_switch==1 and pl_5_switch==1 and pl_6_switch==1 and pl_7_switch==1:
                    print("pl_1/2/3/4/5/6/7")
                    e13='msgstr_plural=[pl_1, pl_2, pl_3, pl_4, pl_5, pl_6, pl_7]'
                    e_+=c+e7+c+e9+c+e10+c+e11+c+e12+c+e13
                else:
                    e_+=c+e7+c+e9+c+e10+c+e11+c+e12

                

                #msgstr_plural -> only 7 supported
                #write flags as string

                '''
                plen = len(e13)
               

                p_header = 0
                for ppart in range(int(0),int(plen)):
                    p_header = p_header + 1
                    p_pt = entry.e13[ppart]
                    print(str(p_pt))

                '''                


            dyf= eval("polib_a.POEntry(" + e_ + ")")

            
            entry = dyf
            
                
           

     
            po.append(entry)

        
       
        po.save(filenameGL)

        #not a multi line string file
        if escapeswitch==1:
            print("Encoding... ")
            print("  ")
       
            time.sleep(5)
         
            #with open(filenameGL, 'r') as file:
            #with open(filenameGL, encoding="utf8") as file:
            with open(filenameGL, encoding=encoding_) as file:
                filedata = file.read()
            

           
                filedata = filedata.replace(r'\\r', r'\r')
                filedata = filedata.replace(r'\\n', r'\n')
       
                '''
                if literalonly==0 and os_switch==1:
                    with open(filenameGL, 'w', newline='\n', encoding=encoding_) as file:
                      file.write(filedata)
                      
                if literalonly==0 and os_switch==0:
                    with open(filenameGL, 'w', newline='\r\n', encoding=encoding_) as file:
                      file.write(filedata)
                '''
                
                if literalonly==1 and ms_switch==0:
                    with open(filenameGL, 'w', newline='\n', encoding=encoding_) as file:
                      file.write(filedata)
                
                # newline='\r\n',
                if literalonly==1 and ms_switch==1:
                    with open(filenameGL, 'w', newline='\r\n', encoding=encoding_) as file:
                      file.write(filedata)
 
            
        #a multi line string file
        if escapeswitch==0:
            print("Encoding multi line string file... ")
            print("  ")
       
            time.sleep(5)
         
            #with open(filenameGL, 'r') as file:
            #with open(filenameGL, encoding="utf8") as file:
            with open(filenameGL, encoding=encoding_) as file:
                filedata = file.read()
            

           
                filedata = filedata.replace(r'\\r', '\r')
                filedata = filedata.replace(r'\\n', '\n')
       
                '''
                if literalonly==0 and os_switch==1:
                    with open(filenameGL, 'w', newline='\n', encoding=encoding_) as file:
                      file.write(filedata)
                      
                if literalonly==0 and os_switch==0:
                    with open(filenameGL, 'w', newline='\r\n', encoding=encoding_) as file:
                      file.write(filedata)
                '''
                
                if literalonly==1 and ms_switch==0:
                    with open(filenameGL, 'w', newline='\n', encoding=encoding_) as file:
                      file.write(filedata)
                
                #newline='\r\n',
                if literalonly==1 and ms_switch==1:
                    with open(filenameGL, 'w', newline='\r\n', encoding=encoding_) as file:
                      file.write(filedata)  
        





        sys.stdout.write(filenameGL)
        
        sys.stdout.write(" ... Done.")
        sys.stdout.flush()        

        mainloop()       

        
        
    def Manual(self):

          

        master = Tk()

        w = Label(master, text="\n\n    Licenses.txt created. The file is in the gnu_gettext_po_tool folder.    \n\n", bg="green")
        w.pack()
        
        with open("Licenses.txt",'w') as gcman:
            
            gcman.write("Licenses")
            gcman.write("\n\n")
            
            gcman.write(OPLH)
            gcman.write("\n")            
            gcman.write(OPLT)

            gcman.write("\n\n")
            gcman.write(JDCALH)
            gcman.write("\n")            
            gcman.write(JDCALT)        

            gcman.write("\n\n")
            gcman.write(ETXMLLH)
            gcman.write("\n")            
            gcman.write(ETXMLLT)
            
            gcman.write("\n\n")
            gcman.write(POLIBLH)
            gcman.write("\n")            
            gcman.write(POLIBLT)


            
         


        mainloop()






def main():

  


    root = Tk()
    ex = xlMain(root)
    #root.geometry("700x180+500+500")
    root.geometry("400x250")
    root.mainloop()  


if __name__ == '__main__':
    main()
